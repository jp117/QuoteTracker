import openpyxl
import os
import datetime

workbook = openpyxl.load_workbook('QuoteTracking.xlsx')
activeQuotes = workbook['ActiveQuotes']

def nextrow(sheet):
    i = 1
    row = sheet.cell(row=i, column=4).value
    while row != None:
        i += 1
        row = sheet.cell(row=i, column=4).value
    return i

def quoteData(sheet, salesman):
    datalistoflist = []
    i = 2
    for i in range (i, nextrow(sheet)):
        data = sheet.cell(row=i, column=4).value
        if data.lower() == salesman.lower() and (sheet.cell(row=i, column=7).value == None or sheet.cell(row=i, column=7).value < datetime.datetime.now()):
            datalist = []
            for x in range (1,8):
                datalist.append(sheet.cell(row=i, column=x).value)
            datalistoflist.append(datalist)            
    return datalistoflist

def subbodydefault(salesman):
    i = len(quoteData(activeQuotes, salesman))
    subbody = ""
    for x in range (0,i):
        if quoteData(activeQuotes, salesman)[x][0] != None:
            subbody += "<b> Quote # " + str(quoteData(activeQuotes, salesman)[x][0]) + "</b> "
        if quoteData(activeQuotes, salesman)[x][1] != None and quoteData(activeQuotes, salesman)[x][2] != None:
            subbody += str(quoteData(activeQuotes, salesman)[x][1]) + " " + str(quoteData(activeQuotes, salesman)[x][2]) + "<br />"
        elif quoteData(activeQuotes, salesman)[x][1] != None:
            subbody += str(quoteData(activeQuotes, salesman)[x][1]) + "<br />"
        else:
            subbody += str(quoteData(activeQuotes, salesman)[x][2]) + "<br />"
        subbody += "Amount Quoted: $ " + str(format(quoteData(activeQuotes, salesman)[x][4], ',.2f')) + "<br />"
        subbody += "Customers Quoted: " + str(quoteData(activeQuotes, salesman)[x][5]) + "<br />"
        subbody += "<br />"
    return subbody

def emailSubBodySalesman(salesman):
    subbody = """\
    <html>
        <head>
        </head>
        <body>
            <p><b>This is a list of quotes that you should follow up on this week</b><br />
            Please remember to send John any quotes you did over 100k and who you sent them to so that they can be tracked.<br /><br />
            """
    subbody += subbodydefault(salesman)
    subbody += "</body>"
    return subbody

def emailSubBodyExec(executive):
    subbody = """\
    <html>
        <head>
        </head>
        <body>
            <p><b>This is a list of quotes that you should follow up on this week</b><br />
            Please remember to send John any quotes you did over 100k and who you sent them to so that they can be tracked.<br /><br />
    """
    subbody += subbodydefault(executive)
    subbody += "<br /> <b>This is what the rest of the team should be following up on:</b><br /><br />Steve<br />"
    subbody += subbodydefault("Steve") + "<br />Harris<br />"
    subbody += subbodydefault("Harris") + "<br />Matt<br />"
    #If new salesman, add here with similar format to above
    subbody += subbodydefault("Matt") + "<br />"
    if executive == "John":
        subbody += "Gina<br />" + subbodydefault("Gina")
    elif executive == "Gina":
        subbody += "John<br />" + subbodydefault("John")
    subbody += "</body>"
    return subbody